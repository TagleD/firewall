import csv
from openpyxl.styles import numbers
import openpyxl
from django.contrib.auth import get_user_model
from django.http import HttpResponse

from webapp.models import Connection, Report


def export_to_excel(request, pk):
    from openpyxl.utils import get_column_letter

    user = get_user_model().objects.get(pk=pk)
    report_id = request.GET.get('report_id')
    is_fraud = request.GET.get('is_fraud')
    connections = Connection.objects.filter(report__user=user)

    if is_fraud == "1":
        connections = connections.filter(is_fraud=True)
    elif is_fraud == "0":
        connections = connections.filter(is_fraud=False)

    if report_id:
        try:
            report = Report.objects.get(id=report_id, user=user)
            connections = connections.filter(report=report)
        except Report.DoesNotExist:
            connections = Connection.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Соединения"

    headers = [
                  "№", "Connection ID", "Malicious", "Date", "Risk (%)", "IP", "Port", "Bytes", "Connection Time (sec)",
                  "Protocol", "Coordinates", "Location", "Rules",
              ] + [f"V{i}" for i in range(1, 29)]
    ws.append(headers)

    for i, t in enumerate(connections, start=1):
        formatted_date = t.time.strftime('%d.%m.%Y %H:%M') if t.time else ''
        ws.append([
            i,
            t.connect_varchar_id,
            "Yes" if t.is_fraud else "No",
            formatted_date,
            t.risk_score,
            t.ip_address,
            t.port,
            t.bytes,
            t.connection_time,
            t.protocol,
            f"{t.latitude}, {t.longitude}" if t.latitude and t.longitude else "",
            t.location,
            t.firewall_rule,
            *[getattr(t, f'v{i}') for i in range(1, 29)],
        ])

    # Автоширина столбцов
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="connections.xlsx"'
    wb.save(response)
    return response


def export_to_csv(request, pk):
    user = get_user_model().objects.get(pk=pk)

    report_id = request.GET.get('report_id')
    is_fraud = request.GET.get('is_fraud')
    connections = Connection.objects.filter(report__user=user)

    if is_fraud == "1":
        connections = connections.filter(is_fraud=True)
    elif is_fraud == "0":
        connections = connections.filter(is_fraud=False)

    if report_id:
        try:
            report = Report.objects.get(id=report_id, user=user)
            connections = connections.filter(report=report)
        except Report.DoesNotExist:
            connections = Connection.objects.none()

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="connections.csv"'

    writer = csv.writer(response)
    headers = [
                  "№", "Connection ID", "Malicious", "Date", "Risk (%)", "IP", "Port", "Bytes", "Connection Time (sec)",
                  "Protocol", "Coordinates", "Location", "Rules",
              ] + [f"V{i}" for i in range(1, 29)]
    writer.writerow(headers)

    for i, t in enumerate(connections, start=1):
        formatted_date = t.time.strftime('%d.%m.%Y %H:%M') if t.time else ''
        coordinates = f"{t.latitude}, {t.longitude}" if t.latitude and t.longitude else ""
        writer.writerow([
            i,
            t.connect_varchar_id,
            "Yes" if t.is_fraud else "No",
            formatted_date,
            t.risk_score,
            t.ip_address or "",
            t.port or "",
            t.bytes or "",
            t.connection_time or "",
            t.protocol or "",
            coordinates,
            t.location or "",
            t.firewall_rule or "",
            *[getattr(t, f'v{i}') for i in range(1, 29)],
        ])

    return response
